"""
Report Generation Service

Generates comprehensive analytics reports in multiple formats:
- PDF reports with charts
- Excel spreadsheets
- CSV exports
"""

import asyncio
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import io
import json
from collections import defaultdict
import statistics

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..config.database import get_db_pool
from ..models.enhanced_models import AnalyticsReport, ReportSection
from ..utils.logger import get_logger

logger = get_logger(__name__)


class ReportGenerationService:
    """
    Service for generating analytics reports
    
    Features:
    - Student performance summary reports
    - Class analytics reports
    - Institution-wide overview reports
    - Multiple export formats (PDF, Excel, CSV)
    """
    
    def __init__(self):
        self.db_pool = None
        self.styles = None
        
    async def initialize(self):
        """Initialize service"""
        try:
            self.db_pool = await get_db_pool()
            if REPORTLAB_AVAILABLE:
                self.styles = getSampleStyleSheet()
            logger.info("ReportGenerationService initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize ReportGenerationService: {e}")
            raise
    
    async def generate_student_report(
        self,
        student_id: str,
        course_id: Optional[str] = None,
        format: str = "pdf"
    ) -> bytes:
        """
        Generate comprehensive student performance report
        
        Args:
            student_id: Student identifier
            course_id: Optional course filter
            format: Output format ("pdf", "excel", "csv")
            
        Returns:
            Report file as bytes
        """
        try:
            logger.info(f"Generating student report for {student_id} in format {format}")
            
            # Collect student data
            student_data = await self._get_student_data(student_id, course_id)
            
            # Build report sections
            sections = []
            
            #Section 1: Overview
            sections.append(await self._build_overview_section(student_data))
            
            # Section 2: Performance Summary
            sections.append(await self._build_performance_section(student_data))
            
            # Section 3: Exam Details
            sections.append(await self._build_exam_details_section(student_data))
            
            # Section 4: Strengths & Weaknesses
            sections.append(await self._build_strengths_section(student_data))
            
            # Section 5: Recommendations
            sections.append(await self._build_recommendations_section(student_data))
            
            # Generate report in requested format
            if format.lower() == "pdf":
                return await self._generate_pdf_report(student_id, sections, "Student Performance Report")
            elif format.lower() == "excel":
                return await self._generate_excel_report(student_id, sections, "Student Performance Report")
            elif format.lower() == "csv":
                return await self._generate_csv_report(student_data)
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.error(f"Error generating student report: {e}")
            raise
    
    async def generate_class_report(
        self,
        course_id: str,
        exam_id: Optional[str] = None,
        format: str = "pdf"
    ) -> bytes:
        """
        Generate class analytics report
        
        Includes:
        - Class performance summary
        - Score distribution
        - Item analysis
        - Student comparisons
        """
        try:
            logger.info(f"Generating class report for course {course_id}")
            
            # Collect class data
            class_data = await self._get_class_data(course_id, exam_id)
            
            # Build sections
            sections = []
            
            sections.append(await self._build_class_overview_section(class_data))
            sections.append(await self._build_score_distribution_section(class_data))
            sections.append(await self._build_class_insights_section(class_data))
            sections.append(await self._build_at_risk_students_section(class_data))
            
            # Generate report
            if format.lower() == "pdf":
                return await self._generate_pdf_report(course_id, sections, "Class Analytics Report")
            elif format.lower() == "excel":
                return await self._generate_excel_report(course_id, sections, "Class Analytics Report")
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.error(f"Error generating class report: {e}")
            raise
    
    async def generate_institution_report(
        self,
        institution_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format: str = "pdf"
    ) -> bytes:
        """
        Generate institution-wide analytics report
        
        Includes:
        - Overall performance trends
        - Course comparisons
        - Faculty effectiveness
        - Student outcomes
        """
        try:
            logger.info(f"Generating institution report for {institution_id}")
            
            # Set date range
            if not end_date:
                end_date = datetime.utcnow()
            if not start_date:
                start_date = end_date - timedelta(days=90)  # Last 90 days
            
            # Collect institution data
            institution_data = await self._get_institution_data(
                institution_id, start_date, end_date
            )
            
            # Build sections
            sections = []
            
            sections.append(await self._build_institution_overview_section(institution_data))
            sections.append(await self._build_course_comparison_section(institution_data))
            sections.append(await self._build_trends_section(institution_data))
            
            # Generate report
            if format.lower() == "pdf":
                return await self._generate_pdf_report(
                    institution_id, sections, "Institution Analytics Report"
                )
            elif format.lower() == "excel":
                return await self._generate_excel_report(
                    institution_id, sections, "Institution Analytics Report"
                )
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.error(f"Error generating institution report: {e}")
            raise
    
    # ========================================================================
    # DATA COLLECTION METHODS
    # ========================================================================
    
    async def _get_student_data(self, student_id: str, course_id: Optional[str]) -> Dict:
        """Collect comprehensive student data"""
        async with self.db_pool.acquire() as conn:
            # Get student info
            student_query = """
                SELECT id, name, email, created_at
                FROM users
                WHERE id = $1
            """
            student = await conn.fetchrow(student_query, student_id)
            
            # Get exam attempts
            exam_query = """
                SELECT 
                    e.id, e.title, e.course_id, e.max_score,
                    a.score, a.completed_at, a.time_taken_minutes
                FROM exam_attempts a
                JOIN exams e ON a.exam_id = e.id
                WHERE a.student_id = $1 AND a.status = 'completed'
            """
            if course_id:
                exam_query += " AND e.course_id = $2"
                exams = await conn.fetch(exam_query, student_id, course_id)
            else:
                exams = await conn.fetch(exam_query, student_id)
            
            return {
                'student': dict(student) if student else {},
                'exams': [dict(e) for e in exams],
                'total_exams': len(exams),
                'avg_score': statistics.mean([e['score'] for e in exams]) if exams else 0,
                'course_id': course_id
            }
    
    async def _get_class_data(self, course_id: str, exam_id: Optional[str]) -> Dict:
        """Collect class performance data"""
        async with self.db_pool.acquire() as conn:
            # Get course info
            course_query = "SELECT id, name, code FROM courses WHERE id = $1"
            course = await conn.fetchrow(course_query, course_id)
            
            # Get all student scores
            if exam_id:
                scores_query = """
                    SELECT student_id, score, max_score, completed_at
                    FROM exam_attempts
                    WHERE exam_id = $1 AND status = 'completed'
                """
                scores = await conn.fetch(scores_query, exam_id)
            else:
                scores_query = """
                    SELECT a.student_id, a.score, e.max_score, a.completed_at
                    FROM exam_attempts a
                    JOIN exams e ON a.exam_id = e.id
                    WHERE e.course_id = $1 AND a.status = 'completed'
                """
                scores = await conn.fetch(scores_query, course_id)
            
            score_values = [s['score'] for s in scores]
            
            return {
                'course': dict(course) if course else {},
                'exam_id': exam_id,
                'scores': [dict(s) for s in scores],
                'total_students': len(set(s['student_id'] for s in scores)),
                'total_attempts': len(scores),
                'avg_score': statistics.mean(score_values) if score_values else 0,
                'min_score': min(score_values) if score_values else 0,
                'max_score': max(score_values) if score_values else 0,
                'median_score': statistics.median(score_values) if score_values else 0
            }
    
    async def _get_institution_data(
        self,
        institution_id: str,
        start_date: datetime,
        end_date: datetime
    ) -> Dict:
        """Collect institution-wide data"""
        async with self.db_pool.acquire() as conn:
            # Get aggregate statistics
            query = """
                SELECT 
                    COUNT(DISTINCT a.student_id) as total_students,
                    COUNT(DISTINCT e.course_id) as total_courses,
                    COUNT(DISTINCT a.id) as total_attempts,
                    AVG(a.score) as avg_score
                FROM exam_attempts a
                JOIN exams e ON a.exam_id = e.id
                WHERE a.status = 'completed'
                  AND a.completed_at BETWEEN $1 AND $2
            """
            stats = await conn.fetchrow(query, start_date, end_date)
            
            return {
                'institution_id': institution_id,
                'start_date': start_date,
                'end_date': end_date,
                'stats': dict(stats) if stats else {}
            }
    
    # ========================================================================
    # SECTION BUILDERS
    # ========================================================================
    
    async def _build_overview_section(self, student_data: Dict) -> ReportSection:
        """Build student overview section"""
        student = student_data['student']
        
        content = {
            'student_name': student.get('name', 'Unknown'),
            'student_id': student.get('id', 'Unknown'),
            'total_exams_completed': student_data['total_exams'],
            'overall_average': f"{student_data['avg_score']:.2f}%",
            'report_date': datetime.utcnow().strftime("%Y-%m-%d")
        }
        
        insights = [
            f"Student has completed {student_data['total_exams']} exams",
            f"Overall average score: {student_data['avg_score']:.1f}%"
        ]
        
        return ReportSection(
            section_title="Student Overview",
            section_type="summary",
            content=content,
            insights=insights
        )
    
    async def _build_performance_section(self, student_data: Dict) -> ReportSection:
        """Build performance summary section"""
        exams = student_data['exams']
        
        if exams:
            scores = [e['score'] for e in exams]
            content = {
                'average_score': statistics.mean(scores),
                'highest_score': max(scores),
                'lowest_score': min(scores),
                'score_range': max(scores) - min(scores),
                'consistency': 'High' if statistics.stdev(scores) < 10 else 'Medium' if statistics.stdev(scores) < 20 else 'Low'
            }
        else:
            content = {'message': 'No exam data available'}
        
        return ReportSection(
            section_title="Performance Summary",
            section_type="table",
            content=content,
            insights=[]
        )
    
    async def _build_exam_details_section(self, student_data: Dict) -> ReportSection:
        """Build exam details table"""
        exams = student_data['exams']
        
        exam_list = []
        for exam in exams:
            exam_list.append({
                'exam_title': exam['title'],
                'score': f"{exam['score']:.1f}",
                'max_score': exam['max_score'],
                'percentage': f"{(exam['score']/exam['max_score']*100):.1f}%" if exam['max_score'] > 0 else "N/A",
                'date': exam['completed_at'].strftime("%Y-%m-%d") if exam['completed_at'] else "N/A"
            })
        
        return ReportSection(
            section_title="Exam Details",
            section_type="table",
            content={'exams': exam_list},
            insights=[]
        )
    
    async def _build_strengths_section(self, student_data: Dict) -> ReportSection:
        """Build strengths and weaknesses section"""
        exams = student_data['exams']
        
        if exams:
            scores = [e['score'] for e in exams]
            avg = statistics.mean(scores)
            
            strengths = []
            weaknesses = []
            
            if avg >= 80:
                strengths.append("Consistent high performance")
            if len(exams) > 5:
                strengths.append("Good course engagement")
            
            if avg < 60:
                weaknesses.append("Below average performance - needs improvement")
            if statistics.stdev(scores) > 20:
                weaknesses.append("Inconsistent scores - focus on study habits")
        else:
            strengths = []
            weaknesses = ["Insufficient data"]
        
        return ReportSection(
            section_title="Strengths & Areas for Improvement",
            section_type="text",
            content={'strengths': strengths, 'weaknesses': weaknesses},
            insights=[]
        )
    
    async def _build_recommendations_section(self, student_data: Dict) -> ReportSection:
        """Build recommendations section"""
        recommendations = [
            "Continue attending all class sessions",
            "Review exam feedback carefully",
            "Practice with additional problems"
        ]
        
        if student_data['avg_score'] < 70:
            recommendations.append("Consider meeting with instructor for extra help")
            recommendations.append("Join study group for peer support")
        
        return ReportSection(
            section_title="Recommendations",
            section_type="text",
            content={'recommendations': recommendations},
            insights=[]
        )
    
    async def _build_class_overview_section(self, class_data: Dict) -> ReportSection:
        """Build class overview section"""
        content = {
            'course_name': class_data['course'].get('name', 'Unknown'),
            'total_students': class_data['total_students'],
            'total_attempts': class_data['total_attempts'],
            'class_average': f"{class_data['avg_score']:.2f}%",
            'median_score': f"{class_data['median_score']:.2f}%",
            'score_range': f"{class_data['min_score']:.1f} - {class_data['max_score']:.1f}"
        }
        
        return ReportSection(
            section_title="Class Overview",
            section_type="summary",
            content=content,
            insights=[]
        )
    
    async def _build_score_distribution_section(self, class_data: Dict) -> ReportSection:
        """Build score distribution section"""
        scores = [s['score'] for s in class_data['scores']]
        
        # Create distribution buckets
        buckets = defaultdict(int)
        for score in scores:
            if score >= 90:
                buckets['A (90-100)'] += 1
            elif score >= 80:
                buckets['B (80-89)'] += 1
            elif score >= 70:
                buckets['C (70-79)'] += 1
            elif score >= 60:
                buckets['D (60-69)'] += 1
            else:
                buckets['F (<60)'] += 1
        
        return ReportSection(
            section_title="Score Distribution",
            section_type="chart",
            content={'distribution': dict(buckets)},
            insights=[]
        )
    
    async def _build_class_insights_section(self, class_data: Dict) -> ReportSection:
        """Build class insights"""
        avg = class_data['avg_score']
        
        insights = []
        if avg >= 80:
            insights.append("Class is performing well overall")
        elif avg < 60:
            insights.append("Class average is below expectations - intervention needed")
        
        return ReportSection(
            section_title="Insights",
            section_type="text",
            content={'insights': insights},
            insights=insights
        )
    
    async def _build_at_risk_students_section(self, class_data: Dict) -> ReportSection:
        """Build at-risk students section"""
        # Simple at-risk detection based on scores
        at_risk = []
        for score in class_data['scores']:
            if score['score'] < 60:
                at_risk.append(score['student_id'])
        
        return ReportSection(
            section_title="At-Risk Students",
            section_type="text",
            content={'at_risk_count': len(set(at_risk))},
            insights=[f"{len(set(at_risk))} students may be at risk"]
        )
    
    async def _build_institution_overview_section(self, data: Dict) -> ReportSection:
        """Build institution overview"""
        stats = data['stats']
        
        content = {
            'total_students': stats.get('total_students', 0),
            'total_courses': stats.get('total_courses', 0),
            'total_exams': stats.get('total_attempts', 0),
            'institution_average': f"{stats.get('avg_score', 0):.2f}%"
        }
        
        return ReportSection(
            section_title="Institution Overview",
            section_type="summary",
            content=content,
            insights=[]
        )
    
    async def _build_course_comparison_section(self, data: Dict) -> ReportSection:
        """Build course comparison section"""
        return ReportSection(
            section_title="Course Performance Comparison",
            section_type="chart",
            content={'courses': []},
            insights=[]
        )
    
    async def _build_trends_section(self, data: Dict) -> ReportSection:
        """Build trends section"""
        date_range = f"{data['start_date'].strftime('%Y-%m-%d')} to {data['end_date'].strftime('%Y-%m-%d')}"
        
        return ReportSection(
            section_title="Performance Trends",
            section_type="text",
            content={'date_range': date_range},
            insights=[]
        )
    
    # ========================================================================
    # REPORT GENERATORS
    # ========================================================================
    
    async def _generate_pdf_report(
        self,
        entity_id: str,
        sections: List[ReportSection],
        report_title: str
    ) -> bytes:
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("ReportLab not installed. Install with: pip install reportlab")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            
            # Add title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a73e8'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            story.append(Paragraph(report_title, title_style))
            story.append(Spacer(1, 0.5 * inch))
            
            # Add sections
            for section in sections:
                # Section title
                story.append(Paragraph(section.section_title, self.styles['Heading2']))
                story.append(Spacer(1, 0.2 * inch))
                
                # Section content
                if section.section_type == "table":
                    if isinstance(section.content, dict):
                        # Convert dict to table
                        data = [[k, str(v)] for k, v in section.content.items()]
                        t = Table(data, colWidths=[3 * inch, 3 * inch])
                        t.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 14),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(t)
                else:
                    # Add text content
                    content_text = json.dumps(section.content, indent=2)
                    story.append(Paragraph(content_text, self.styles['BodyText']))
                
                story.append(Spacer(1, 0.3 * inch))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            raise
    
    async def _generate_excel_report(
        self,
        entity_id: str,
        sections: List[ReportSection],
        report_title: str
    ) -> bytes:
        """Generate Excel report"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Add title
            ws['A1'] = report_title
            ws['A1'].font = Font(size=18, bold=True)
            ws.merge_cells('A1:D1')
            
            row = 3
            
            # Add sections
            for section in sections:
                # Section title
                ws[f'A{row}'] = section.section_title
                ws[f'A{row}'].font = Font(size=14, bold=True)
                row += 1
                
                # Section content
                if isinstance(section.content, dict):
                    for key, value in section.content.items():
                        ws[f'A{row}'] = key
                        ws[f'B{row}'] = str(value)
                        row += 1
                
                row += 2  # Space between sections
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            raise
    
    async def _generate_csv_report(self, student_data: Dict) -> bytes:
        """Generate simple CSV report"""
        try:
            buffer = io.StringIO()
            
            # Write header
            buffer.write("Exam Title,Score,Max Score,Percentage,Date\n")
            
            # Write exam data
            for exam in student_data['exams']:
                percentage = (exam['score'] / exam['max_score'] * 100) if exam['max_score'] > 0 else 0
                date_str = exam['completed_at'].strftime("%Y-%m-%d") if exam['completed_at'] else "N/A"
                buffer.write(
                    f"{exam['title']},{exam['score']},{exam['max_score']},{percentage:.2f}%,{date_str}\n"
                )
            
            return buffer.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")
            raise


# Global service instance
report_service = ReportGenerationService()


async def get_report_service() -> ReportGenerationService:
    """Get the report generation service instance"""
    return report_service
